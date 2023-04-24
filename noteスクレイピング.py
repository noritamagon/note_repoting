#スクレイピングの定義関数#########################################
def scraping():

  from selenium import webdriver
  from time import sleep

  #ログイン画面の操作
  browser=webdriver.Chrome("C:/Users/名前/chromedriver.exe")

  note_url=browser.get("https://note.com/login")
  
  user_id=browser.find_element_by_name("login")
  user_id.send_keys("メールアドレス")

  user_pass=browser.find_element_by_name("password")
  user_pass.send_keys("パスワード")

  login_button=browser.find_element_by_class_name("logining_msg")
  login_button.click()

  sleep(5)

  #ダッシュボード移動
　dash_board=browser.get("https://note.com/sitesettings/stats")
  
  #週ごとのデータに切り替える
  week_button=browser.find_element_by_class_name("btn.btn--size_small")
  week_button.click()
  sleep(5)

  #データの取得
  #PV数の取得
  PV_num=browser.find_element_by_class_name("0-statsContent__overviewItem.o-statsContent__overviewItem--type_view")
  PV_dic={}
  PV_list=[]
  PV_list=PV_num.text.split("\n")
  PV_dic[PV_list[1]]=PV_list[0]

  #スキの数取得
  PV_num2=browser.find_element_by_class_name("0-statsContent__overviewItem.o-statsContent__overviewItem--type_suki")
  PV_list=PV_num2.text_split("\n")
  PV_dic[PV_list[1]]==PV_list[0]

  sleep(5)

  #フォロワー数の取得
  follower=browser.get("https://note.com/ユーザー名")
  fol_num=browser.find_element_by_class_name("m-profileStatus__follow")

  fol_list=[]
  fol_list2=[]
  fol_list=fol_num.text.split("\n")
  fol_list2=fpl_list[1].split(" ")

  PV_dic[fol_list2[1]]=fol_list2[0]

  #データフレームの定義
  import pandas as pd

  df=pd.DataFrame(PV_dic.values(),index=PV_dic.keys()).T

  import datetime

  now=datetime.datetime.today().strftime("%Y{}%m{}%d{}").format(*"年月日")

  df["日次"]=now

  df.to_csv("noteスクレイピング"+"{}.csv".format(now),encoding="shift-jis",index=False)

  browser.quit()


#データ貯蓄用Excelシートに取得結果を転記する#########################################
#Excelファイルは事前に用意
def totalization():
  import openpyxl
  wb=openpyxl.load_workbook("noteデータ貯蓄.xlsx")
  ws=wb["Sheet1"]

  import datetime
  import pandas as pd
  
  #保存したファイルを呼び出して、貯蓄シートに追加する
  now=datetime.datetime.today().strftime("%Y{}%m{}%d{}").format(*"年月日")
  df=pd.read_csv("noteスクレイピング"+"{}.csv".format(now),encoding="shift-jis")
  now_list=df.iloc[0].tolist()

  now_list2=[]
  for i in range(0,len(now_list)-1):
    now_list2.append(now_list[i].item())

  now=now_list[3]
  import locale
  locale.setlocale(locale.LC_CTYPE,"Japanese_Japan.932")

  now=datetime.datetime.strptime(now,"%Y年%m月%d日")
  now=now.strftime("%Y年%m月%d日")
  now_list2.append(now)

  ws.append(now_list2)

  #最終行以降に新規データを追加
  masRow=ws.max_row

  for j in range(5,8):
    i=maxRow
    try:
      ws.cell(i,j).value=ws.cell(i,j-4).value/ws.cell(i-1,j-4).value
      ws.cell(i,j).number_format="0%"
    except: ZeroDivisionError:
      ws.cell(i,j).value=0
      es.cell(i,j).number_format="0%"

  wb.save("noteデータ貯蓄.xlsx") #上書き保存


#Excel内のグラフの更新#########################################
def LineChart():
  import openpyxl
  from openpyxl import Workbook
  from openpyxl.chart import LineChart,Reference

  wb=openpyxl.load_workbook("noteデータ貯蓄.xlsx")
  ws=wb["Sheet1"]

  #グラフの定義
  maxRow=ws.max_row
  graph_title=[]

  #グラフタイトルをカラム名から抽出
  for j in range(0,3):
    titles=ws.cell(1,j+5).value
    graph_title.append(titles)

  #PV数、フォロー数、良い数の前週比のグラフをそれぞれ作成
  for i in range(5,8):

    values=Reference(ws,min_col=i,min_row=2,max_col=i,max_row=maxRow)      #値のセット
    categories=Reference(ws,min_col=4,min_row=2,max_col=4,max_row=maxRow)  #横軸のラベル

    #折れ線グラフの作成
    chart=LineChart()
    graph_title_number=i-5
    chart.title=graph_title[graph_title_number] #グラフタイトルをグラフに設定

    chart.add_data(values)    #グラフに値を設定
    chart.set_categories(categories) #グラフに横軸を設定
    chart.legend=None                #凡例は不要

    output_cell=12*(i-4)

    ws.add_chart(chart,"1{}".format(output_cell)) #セル番地を指定してグラフを作成

  wb.save("noteデータ貯蓄.xlsx") #上書き保存


#各処理を実行#########################################
scraping()
totalization()
LineChart()  